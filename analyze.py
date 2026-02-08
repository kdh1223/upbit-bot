import csv
import os
import time
from dataclasses import dataclass, asdict
from datetime import datetime
from typing import Dict, List, Tuple

import config


TRADE_LOG = getattr(config, "TRADE_LOG_PATH", "trade_log.csv")
OUT_XLSX = "trading_summary.xlsx"
OUT_SUMMARY_CSV = "trading_summary.csv"

_LAST_RUN_AT = 0.0


def _to_float(x, default=0.0) -> float:
    try:
        return float(x)
    except Exception:
        return default


def _get_initial_capital() -> float:
    try:
        return float(getattr(config, "INITIAL_CAPITAL", 1_000_000))
    except Exception:
        return 1_000_000.0


def build_equity_curve(header: List[str], rows: List[List[str]]):
    if not header or not rows:
        return [], 0.0, "", ""

    col = {name: idx for idx, name in enumerate(header)}
    idx_time = col.get("time")
    idx_pnl = col.get("pnl_pct")
    if idx_pnl is None:
        return [], 0.0, "", ""

    cost_pp = float(getattr(config, "COST_ROUNDTRIP_PCT", 0.0)) * 100.0
    equity = _get_initial_capital()
    peak = equity
    peak_time = ""
    max_dd = 0.0
    mdd_start = ""
    mdd_end = ""
    curve = []

    for row in rows:
        if idx_pnl >= len(row):
            continue

        ts = row[idx_time] if (idx_time is not None and idx_time < len(row)) else ""
        pnl = _to_float(row[idx_pnl], 0.0)
        pnl_net = pnl - cost_pp
        equity *= (1.0 + pnl_net / 100.0)

        if equity > peak:
            peak = equity
            peak_time = ts

        if peak > 0:
            dd = (equity / peak - 1.0) * 100.0
        else:
            dd = 0.0

        if dd < max_dd:
            max_dd = dd
            mdd_start = peak_time if peak_time else ts
            mdd_end = ts

        curve.append((ts, equity, dd))

    mdd = abs(max_dd)
    return curve, mdd, mdd_start, mdd_end


def load_trades(path: str) -> Tuple[List[str], List[List[str]]]:
    if not os.path.exists(path):
        return [], []
    with open(path, "r", encoding="utf-8") as f:
        r = csv.reader(f)
        rows = list(r)
    if not rows:
        return [], []
    header = rows[0]
    data = rows[1:]
    return header, data


@dataclass
class Metrics:
    generated_at: str
    total: int = 0
    wins: int = 0
    losses: int = 0
    winrate: float = 0.0

    avg_win: float = 0.0
    avg_loss: float = 0.0  # negative
    rr: float = 0.0
    expectancy: float = 0.0  # %p per trade

    gross_compound: float = 0.0  # %
    net_compound: float = 0.0    # %
    cost_pp: float = 0.0         # roundtrip cost in %p
    mdd: float = 0.0             # %
    mdd_start: str = ""
    mdd_end: str = ""
    profit_factor: float = 0.0

    by_reason: Dict[str, Dict[str, float]] = None
    by_regime: Dict[str, Dict[str, float]] = None
    by_strategy: Dict[str, Dict[str, float]] = None


def _summarize_group(arr: List[float]) -> Dict[str, float]:
    if not arr:
        return {"count": 0, "winrate": 0.0, "avg": 0.0}
    winrate = sum(1 for x in arr if x > 0) / len(arr) * 100.0
    avg = sum(arr) / len(arr)
    return {"count": float(len(arr)), "winrate": winrate, "avg": avg}


def calc_metrics(header: List[str], rows: List[List[str]]) -> Metrics:
    cost_pp = float(getattr(config, "COST_ROUNDTRIP_PCT", 0.0)) * 100.0
    m = Metrics(
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        cost_pp=cost_pp,
        by_reason={},
        by_regime={},
        by_strategy={},
    )

    if not header or not rows:
        return m

    col = {name: idx for idx, name in enumerate(header)}
    idx_pnl = col.get("pnl_pct")
    idx_reason = col.get("reason")
    idx_regime = col.get("regime")
    idx_strategy = col.get("strategy")

    pnls: List[float] = []
    pnls_net: List[float] = []

    reason_map: Dict[str, List[float]] = {}
    regime_map: Dict[str, List[float]] = {}
    strategy_map: Dict[str, List[float]] = {}

    for row in rows:
        if idx_pnl is None or idx_pnl >= len(row):
            continue

        pnl = _to_float(row[idx_pnl], 0.0)   # % 단위
        pnls.append(pnl)
        pnls_net.append(pnl - cost_pp)

        reason = row[idx_reason] if idx_reason is not None and idx_reason < len(row) else ""
        regime = row[idx_regime] if idx_regime is not None and idx_regime < len(row) else ""
        strategy = row[idx_strategy] if idx_strategy is not None and idx_strategy < len(row) else ""

        reason_map.setdefault(reason, []).append(pnl)
        regime_map.setdefault(regime, []).append(pnl)
        strategy_map.setdefault(strategy, []).append(pnl)

    m.total = len(pnls)
    if m.total == 0:
        return m

    m.wins = sum(1 for x in pnls if x > 0)
    m.losses = sum(1 for x in pnls if x < 0)
    m.winrate = (m.wins / m.total) * 100.0

    win_list = [x for x in pnls if x > 0]
    loss_list = [x for x in pnls if x < 0]

    m.avg_win = (sum(win_list) / len(win_list)) if win_list else 0.0
    m.avg_loss = (sum(loss_list) / len(loss_list)) if loss_list else 0.0
    m.rr = (m.avg_win / abs(m.avg_loss)) if (m.avg_win > 0 and m.avg_loss < 0) else 0.0

    p = m.wins / m.total
    m.expectancy = p * m.avg_win + (1 - p) * m.avg_loss

    gross_mult = 1.0
    net_mult = 1.0
    for x in pnls:
        gross_mult *= (1.0 + x / 100.0)
    for x in pnls_net:
        net_mult *= (1.0 + x / 100.0)

    m.gross_compound = (gross_mult - 1.0) * 100.0
    m.net_compound = (net_mult - 1.0) * 100.0

    # MDD (gross equity curve 기준)
    _, mdd, mdd_start, mdd_end = build_equity_curve(header, rows)
    m.mdd = mdd
    m.mdd_start = mdd_start
    m.mdd_end = mdd_end

    # Profit Factor
    gross_profit = sum(win_list)
    gross_loss = abs(sum(loss_list))
    m.profit_factor = (gross_profit / gross_loss) if gross_loss > 0 else 0.0

    # 그룹 요약 저장
    for k, arr in reason_map.items():
        name = k if k else "(blank)"
        m.by_reason[name] = _summarize_group(arr)
    for k, arr in regime_map.items():
        name = k if k else "(blank)"
        m.by_regime[name] = _summarize_group(arr)
    for k, arr in strategy_map.items():
        name = k if k else "(blank)"
        m.by_strategy[name] = _summarize_group(arr)

    return m


def print_report(m: Metrics):
    print("\n" + "=" * 44)
    print("[REPORT] Trading Performance Report")
    print("=" * 44)
    print(f"Generated: {m.generated_at}")
    print(f"Roundtrip Cost Assumed: {m.cost_pp:.2f}%p")
    print("-" * 44)

    if m.total == 0:
        print("거래 기록이 없습니다. (trade_log.csv 비어있음)")
        print("=" * 44 + "\n")
        return

    print(f"총 거래수: {m.total}")
    print(f"승/패: {m.wins} / {m.losses}")
    print(f"승률: {m.winrate:.1f}%")
    print(f"평균 수익(승): +{m.avg_win:.2f}%")
    print(f"평균 손실(패): {m.avg_loss:.2f}%")
    print(f"손익비(RR): {m.rr:.2f}")
    print(f"기대값(1회당): {m.expectancy:+.2f}%p")
    print(f"복리 누적(수수료 미반영): {m.gross_compound:+.2f}%")
    print(f"복리 누적(왕복비용 반영): {m.net_compound:+.2f}%")
    print("-" * 44)

    if m.by_regime:
        print("[Regime] summary")
        for k, v in sorted(m.by_regime.items(), key=lambda x: -x[1]["count"]):
            print(f"  {k:8s} | n={int(v['count']):3d} | win={v['winrate']:5.1f}% | avg={v['avg']:+.2f}%")

    if m.by_strategy:
        print("\n[Strategy] summary")
        for k, v in sorted(m.by_strategy.items(), key=lambda x: -x[1]["count"]):
            print(f"  {k:8s} | n={int(v['count']):3d} | win={v['winrate']:5.1f}% | avg={v['avg']:+.2f}%")

    if m.by_reason:
        print("\n[Reason] summary")
        for k, v in sorted(m.by_reason.items(), key=lambda x: -x[1]["count"]):
            print(f"  {k:12s} | n={int(v['count']):3d} | win={v['winrate']:5.1f}% | avg={v['avg']:+.2f}%")

    print("=" * 44 + "\n")


def save_summary_csv(m: Metrics, out_path: str):
    rows = [
        ["generated_at", m.generated_at],
        ["total", m.total],
        ["wins", m.wins],
        ["losses", m.losses],
        ["winrate_pct", f"{m.winrate:.2f}"],
        ["avg_win_pct", f"{m.avg_win:.4f}"],
        ["avg_loss_pct", f"{m.avg_loss:.4f}"],
        ["rr", f"{m.rr:.4f}"],
        ["expectancy_pctp", f"{m.expectancy:.4f}"],
        ["gross_compound_pct", f"{m.gross_compound:.4f}"],
        ["net_compound_pct", f"{m.net_compound:.4f}"],
        ["roundtrip_cost_pctp", f"{m.cost_pp:.4f}"],
        ["mdd_pct", f"{m.mdd:.4f}"],
        ["mdd_start", m.mdd_start],
        ["mdd_end", m.mdd_end],
        ["profit_factor", f"{m.profit_factor:.4f}"],
    ]
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["metric", "value"])
        w.writerows(rows)


def save_xlsx(
    header: List[str],
    rows: List[List[str]],
    m: Metrics,
    out_path: str,
    equity_curve=None,
):
    try:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        from openpyxl.utils import get_column_letter
    except Exception:
        print("[WARN] openpyxl missing; skip xlsx (summary CSV is still written)")
        return

    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["metric", "value"])
    for k, v in asdict(m).items():
        if k in ("by_reason", "by_regime", "by_strategy"):
            continue
        ws.append([k, v])

    ws.append([])
    ws.append(["Regime", "count", "winrate_pct", "avg_pct"])
    for k, v in sorted(m.by_regime.items(), key=lambda x: -x[1]["count"]):
        ws.append([k, int(v["count"]), v["winrate"], v["avg"]])

    ws.append([])
    ws.append(["Strategy", "count", "winrate_pct", "avg_pct"])
    for k, v in sorted(m.by_strategy.items(), key=lambda x: -x[1]["count"]):
        ws.append([k, int(v["count"]), v["winrate"], v["avg"]])

    ws.append([])
    ws.append(["Reason", "count", "winrate_pct", "avg_pct"])
    for k, v in sorted(m.by_reason.items(), key=lambda x: -x[1]["count"]):
        ws.append([k, int(v["count"]), v["winrate"], v["avg"]])

    ws2 = wb.create_sheet("Trades")
    if header:
        ws2.append(header)
    for r in rows:
        ws2.append(r)

    sheets = [ws, ws2]

    if equity_curve:
        ws3 = wb.create_sheet("Equity")
        ws3.append(["timestamp", "equity", "drawdown_pct"])
        for ts, eq, dd in equity_curve:
            ws3.append([ts, eq, dd])
        sheets.append(ws3)

        ws4 = wb.create_sheet("Charts")
        sheets.append(ws4)

        max_row = ws3.max_row
        if max_row >= 2:
            cats = Reference(ws3, min_col=1, min_row=2, max_row=max_row)

            eq_chart = LineChart()
            eq_chart.title = "Equity"
            eq_chart.y_axis.title = "Equity"
            eq_chart.x_axis.title = "Time"
            eq_data = Reference(ws3, min_col=2, min_row=1, max_row=max_row)
            eq_chart.add_data(eq_data, titles_from_data=True)
            eq_chart.set_categories(cats)
            ws4.add_chart(eq_chart, "A1")

            dd_chart = LineChart()
            dd_chart.title = "Drawdown"
            dd_chart.y_axis.title = "Drawdown %"
            dd_chart.x_axis.title = "Time"
            dd_data = Reference(ws3, min_col=3, min_row=1, max_row=max_row)
            dd_chart.add_data(dd_data, titles_from_data=True)
            dd_chart.set_categories(cats)
            ws4.add_chart(dd_chart, "A16")

    for wsx in sheets:
        for col_idx in range(1, wsx.max_column + 1):
            letter = get_column_letter(col_idx)
            wsx.column_dimensions[letter].width = 18

    wb.save(out_path)


def generate_report(
    trade_log_path: str = TRADE_LOG,
    out_csv: str = OUT_SUMMARY_CSV,
    out_xlsx: str = OUT_XLSX,
    quiet: bool = False,
):
    header, rows = load_trades(trade_log_path)
    m = calc_metrics(header, rows)
    equity_curve, _, _, _ = build_equity_curve(header, rows)

    if not quiet:
        print_report(m)

    save_summary_csv(m, out_csv)
    save_xlsx(header, rows, m, out_xlsx, equity_curve=equity_curve)
    return m


def maybe_generate_report(
    trade_log_path: str = TRADE_LOG,
    out_csv: str = OUT_SUMMARY_CSV,
    out_xlsx: str = OUT_XLSX,
    min_interval_sec: float = 30.0,
    quiet: bool = True,
):
    global _LAST_RUN_AT
    now = time.time()
    if now - _LAST_RUN_AT < float(min_interval_sec):
        return None
    _LAST_RUN_AT = now
    try:
        return generate_report(trade_log_path, out_csv, out_xlsx, quiet=quiet)
    except PermissionError:
        print(f"[WARN] failed to save {out_xlsx}; file is open. Close it and retry.")
        return None
    except Exception as e:
        try:
            print(f"[WARN] report generation failed: {e}")
        except Exception:
            print(f"[WARN] report generation failed: {e}")
        return None


def main():
    generate_report()

    print(f"[OK] summary CSV saved: {OUT_SUMMARY_CSV}")
    print(f"[OK] xlsx saved: {OUT_XLSX} (if openpyxl installed)")


if __name__ == "__main__":
    main()
